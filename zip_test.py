import os
import csv
import zipfile
from openpyxl import load_workbook
from PyPDF2 import PdfReader

# Проверка, что файлы созданы
def test_files_exist(pdf_file, excel_file, csv_file):
    assert os.path.exists(pdf_file)
    assert os.path.exists(excel_file)
    assert os.path.exists(csv_file)

# Проверка содержимого PDF-файла
def test_pdf_content(pdf_file):
    with open(pdf_file, "rb") as f:
        reader = PdfReader(f)
        assert len(reader.pages) == 1  # Ожидаем 1 страницу

# Проверка содержимого Excel-файла
def test_excel_content(excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    data = list(sheet.iter_rows(values_only=True))
    assert data == [("Name", "Age"), ("Kate", 28), ("Danil", 25)]

# Проверка содержимого CSV-файла
def test_csv_content(csv_file):
    with open(csv_file, newline="") as f:
        reader = csv.reader(f)
        data = list(reader)
    assert data == [["First", "manual"], ["1", "Put the button"], ["2", "Check window"]]

# Проверка содержимого ZIP-архива
def test_zip_content(zip_file):
    with zipfile.ZipFile(zip_file, "r") as archive:
        files = archive.namelist()
    assert set(files) == {"book.pdf", "table.xlsx", "manual.csv"}