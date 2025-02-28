import os  # Модуль для работы с файловой системой
import csv  # Модуль для работы с CSV-файлами
import zipfile  # Модуль для работы с архивами
from openpyxl import Workbook  # Для работы с Excel-файлами
from PyPDF2 import PdfWriter, PdfReader  # Для создания и чтения PDF-файлов


# Создание папки, если ее нет
os.makedirs("tmp", exist_ok=True)

#  Создание PDF-файла
pdf_path = "tmp/book.pdf"
pdf_writer = PdfWriter()
pdf_writer.add_blank_page(width=200, height=200)  # Добавляем пустую страницу в PDF
with open(pdf_path, "wb") as f:  # бинарный режим
    pdf_writer.write(f)

# Создание Excel-файла
xlsx_path = "tmp/table.xlsx"
workbook = Workbook()  # объект Excel
sheet = workbook.active
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet.append(["Kate", 28])
sheet.append(["Danil", 25])
workbook.save(xlsx_path)

# Создание CSV-файла
csv_path = "tmp/manual.csv"
with open(csv_path, mode="w", newline="") as csv_file:
    writer = csv.writer(csv_file)
    writer.writerow(["First", "manual"])  # заголовок
    writer.writerow([1, "Put the button"])
    writer.writerow([2, "Check window"])

# Архивирование файлов
zip_path = "tmp/files.zip"
with zipfile.ZipFile(zip_path, "w") as archive:
    archive.write(pdf_path, arcname="book.pdf")
    archive.write(xlsx_path, arcname="table.xlsx")
    archive.write(csv_path, arcname="manual.csv")

# Чтение содержимого архива без распаковки
with zipfile.ZipFile(zip_path, "r") as archive:  # архив для чтения
    print("Файлы в архиве:", archive.namelist())  # список файлов в архиве
    for file_name in archive.namelist():  # перебираем файлы в архиве
        with archive.open(file_name) as file:  # открываем файл из архива
            print(f"\nСодержимое файла {file_name}:")
            if file_name.endswith(".pdf"):  # Если файл PDF
                with open("tmp/temp_pdf.pdf", "wb") as temp_pdf:  # Временный файл для чтения PDF
                    temp_pdf.write(file.read())
                with open("tmp/temp_pdf.pdf", "rb") as temp_pdf_binary:
                    reader = PdfReader(temp_pdf_binary)
                    print(f"Количество страниц в PDF: {len(reader.pages)}")
            elif file_name.endswith(".xlsx"):  # Если файл Excel
                from openpyxl import load_workbook
                workbook = load_workbook(file)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    print(row)
            elif file_name.endswith(".csv"):  # Если файл CSV
                reader = csv.reader(file.read().decode("utf-8").splitlines())
                for row in reader:
                    print(row)